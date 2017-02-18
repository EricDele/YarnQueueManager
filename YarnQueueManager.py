#!/usr/bin/python
# coding: utf-8

# ===================================================================#
# -------------------------------------------------------------------#
#                         YarnQueueManager                           #
# -------------------------------------------------------------------#
# *******************************************************************#
#                   Eric Deleforterie - 2016/11/21                   #
# -------------------------------------------------------------------#
#                          Notes/Comments                            #
#                                                                    #
# -------------------------------------------------------------------#
#                             HISTORY                                #
#    V0.0.1    Eric Deleforterie - 2016/11/21                        #
#              Creation and first features                           #
# ===================================================================#


# --------------------------------------------#
#             Packages Importation            #
# --------------------------------------------#
from __future__ import print_function
import sys
import re
import json
import pprint
import argparse
import xlsxwriter
import requests
import time
from openpyxl import load_workbook
from lxml import etree
from collections import defaultdict

# --------------------------------------------#
#              Variables declaration          #
# --------------------------------------------#
global vg_fileName
global vg_arguments
global vg_delimiter
global vg_configuration
global vg_configProperties
global vg_xlsConfig
global vg_ambariConfig

DEFAULT = '\033[39m'
BLACK = '\033[30m'
RED = '\033[91m'
GREEN = '\033[92m'
YELLOW = '\033[93m'
BLUE = '\033[94m'
CYAN = '\033[96m'
BACK_RED = '\033[41m'
BACK_GREEN = '\033[42m'
BACK_YELLOW = '\033[43m'
BACK_BLUE = '\033[44m'
BACK_CYAN = '\033[46m'
BACK_GRAY = '\033[47m'
BACK_DEFAULT = '\033[49m'

# --------------------------------------------#
#                 Queue Class                 #
# --------------------------------------------#


class Queues():
    """Object for Queues"""

    def __init__(self, configuration, properties, ambariConfiguration):
        # Queues configuration from xls
        self.queues = defaultdict(dict)
        # Queues configuration from ambari
        self.ambari = defaultdict(dict)
        # General configuration from config file
        self.configuration = configuration
        # Properties list from config file
        self.properties = properties
        # Ambari configuration from config file
        self.ambariConfiguration = ambariConfiguration
        # ADMIN_VIEW version for managing the put mediaType
        self.adminViewVersion = "default"
        # Boolean to know if we could change something or just dry run
        self.dryRun = True

    # --------------------------------------------#
    #   Add a property with its value to a queue  #
    #                                             #
    # --------------------------------------------#

    def addQueueValue(self, arborescence, queueName, propertyName, value):
        # print(propertyName + " : " + str(value))
        try:
            if(type(value) is not dict and propertyName in self.properties):
                if(self.properties[propertyName] == "int"):
                    value = int(value)
                elif(self.properties[propertyName] == "float"):
                    value = float(value)
            if(propertyName == self.configuration['arbo-queues-property']):
                # we have a property for an arborescence with the queues names so
                # we are on a arborescence head
                self.queues[queueName]['arborescence'] = queueName
                self.queues[queueName]['arborescence-head'] = 'yes'
                # Set the queues tree leaf lists from when from XML
                self.queues[queueName][propertyName] = value
            if(arborescence != ""):
                # we have a property for a queue within arborescence
                self.queues['.'.join((arborescence, queueName))][propertyName] = value
                self.queues['.'.join((arborescence, queueName))]['queue-name'] = queueName
                self.queues['.'.join((arborescence, queueName))]['arborescence'] = arborescence
            else:
                # we have a property for a queue that is not in an arborescence
                self.queues[queueName][propertyName] = value
                self.queues[queueName]['queue-name'] = queueName
                self.queues[queueName]['arborescence'] = ''
        except Exception as e:
            raise e

    # --------------------------------------------#
    #      Manage the tree leaf of the Queues     #
    #                                             #
    # --------------------------------------------#

    def manageQueuesTreeLeafs(self):
        for queueName in sorted(self.queues.keys()):
            # Test if we have a multi leaf arborescence
            if queueName.find('.') != -1:
                # Have an arborescence with leaf, split all the leafs in a list
                elements = queueName.split('.')
                nbElements = len(elements)
                # Iterate from the end of the list to obtain the tree leaf
                # WordA.WordB.WordC.WordD : WordE
                # WordA.WordB.WordC : WordD
                # WordA.WordB : WordC
                # WordA : WordB
                for indice in range(nbElements - 1, 0, -1):
                    actualArborescence = '.'.join(elements[0:indice])
                    if 'queues' in self.queues[actualArborescence]:
                        self.queues[actualArborescence]['queues'] = ','.join((self.queues[actualArborescence]['queues'], str(elements[indice])))
                    else:
                        self.queues[actualArborescence]['queues'] = str(elements[indice])
            else:
                # We don't put the root queue in the tree
                if queueName != self.configuration['root-name']:
                    # We are just under root
                    if 'queues' in self.queues[self.configuration['root-name']] and queueName not in self.queues[self.configuration['root-name']]['queues']:
                        self.queues[self.configuration['root-name']]['queues'] = ','.join((self.queues[self.configuration['root-name']]['queues'], queueName))
                    else:
                        self.queues[self.configuration['root-name']]['queues'] = queueName

    # --------------------------------------------#
    #   Compute the capacity for the leafs        #
    #                of a branch                  #
    # --------------------------------------------#

    def computeArborescenceCapacity(self, arborescence):
        arborescenceCapacity = 0
        for queueName in sorted(self.queues.keys()):
            if 'arborescence' in self.queues[queueName] and self.queues[queueName]['arborescence'] == arborescence and 'arborescence-head' not in self.queues[queueName]:
                arborescenceCapacity = arborescenceCapacity + self.queues[queueName]['capacity']
        return arborescenceCapacity

    # --------------------------------------------#
    #   Some checks after reading the Excel file  #
    #                                             #
    # --------------------------------------------#

    def checkQueuesCoherence(self):
        totalCapacity = 0
        checkSuccessfull = True
        for queueName in sorted(self.queues.keys()):
            # Test if maximum-capacity is under capacity
            if 'capacity' in self.queues[queueName] and 'maximum-capacity' in self.queues[queueName]:
                if self.queues[queueName]['capacity'] > self.queues[queueName]['maximum-capacity']:
                    print(BACK_GRAY + RED + "Queue : " + queueName + ", capacity is > maximum-capacity ( " + str(self.queues[queueName]['capacity']) + " > " +
                          str(self.queues[queueName]['maximum-capacity']) + " )" + DEFAULT + BACK_DEFAULT)
                    checkSuccessfull = False
            # Test if we are on a branch with leafs
            if 'capacity' in self.queues[queueName]:
                # A Leaf
                if 'arborescence' in self.queues[queueName] and self.queues[queueName]['arborescence'] == "":
                    totalCapacity = totalCapacity + self.queues[queueName]['capacity']
                # A branch head
                if 'arborescence-head' in self.queues[queueName] and self.queues[queueName]['arborescence-head'] == "yes":
                    # Compute for this branch the totalLeafsCapacity
                    totalLeafsCapacity = self.computeArborescenceCapacity(queueName)
                    if totalLeafsCapacity != 100:
                        checkSuccessfull = False
                        print(BACK_GRAY + RED + "Arborescence Capacity is not 100% for leafs of " + queueName + " : " + str(totalLeafsCapacity) + DEFAULT + BACK_DEFAULT)
        if totalCapacity != 100:
            checkSuccessfull = False
            print(BACK_GRAY + RED + "Total Capacity is not 100% : " + str(totalCapacity) + DEFAULT + BACK_DEFAULT)
        return checkSuccessfull
    # --------------------------------------------#
    #       Create the XLS file from the Queues   #
    #                   object                    #
    # --------------------------------------------#

    def queuesToXLS(self, fileXLS, configXLS):
        print(BACK_GRAY + BLACK + "\nCreating XLS file :" + DEFAULT + BACK_DEFAULT + " " + fileXLS)
        # Create the XLS file
        workbook = xlsxwriter.Workbook(fileXLS)
        # Create the sheet
        worksheet = workbook.add_worksheet(configXLS['sheet-name'])
        # Some paint for the titles
        titleFormat = workbook.add_format({'bold': True, 'font_name': 'calibri', 'font_color': 'white', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#198A8A'})
        worksheet.set_row(int(configXLS['row-titles']), 45, titleFormat)
        # Write column titles
        for columnLetter in sorted(configXLS['topology'].keys()):
            # Check if the column have to appear in the excel file
            if 'columnTitle' in configXLS['topology'][columnLetter]:
                worksheet.write(int(configXLS['row-titles']), self.lettreVersCol(columnLetter), configXLS['topology'][columnLetter]['columnTitle'], titleFormat)
        # Write the root
        worksheet.write(int(configXLS['row-titles']) + 1, 1, str(self.configuration['root-name']))
        ligne = int(configXLS['cellule-origine']['row'])
        column = int(configXLS['cellule-origine']['col'])
        # Revert topology's configuration for property access easier
        revertedConf = self.revertConfigurationDict(configXLS['topology'])
        # Iterat on sorted queues
        for queueName in sorted(self.queues.keys()):
            if(queueName != self.configuration['root-name'] and queueName != self.configuration['general']):

                # TODOOOOOOOOOOO : arborescences management

                # Wrtie queue name
                worksheet.write(ligne, column, queueName)
                # Iterate on column
                for propertyName in self.queues[queueName]:
                    # Check a last one if the property have to be in the XLS file
                    if propertyName in revertedConf:
                        # write the value in the column
                        worksheet.write(ligne, self.lettreVersCol(revertedConf[propertyName]['column']), self.queues[queueName][propertyName])
                ligne += 1
        # XLS file closure
        workbook.close()

    # --------------------------------------------#
    #            save queues to json file         #
    #                                             #
    # --------------------------------------------#
    def queuesToJsonFile(self, file):
        with open(file, 'w') as outfile:
            if self.queues.__len__ > 2:
                json.dump(self.queues, outfile)
            elif self.ambari.__len__ > 2:
                json.dump(self.ambari, outfile)

    # --------------------------------------------#
    #            save config to json file         #
    #                                             #
    # --------------------------------------------#
    def saveQueuesToFile(self, data, file):
        with open(file, 'w') as outfile:
            json.dump(data, outfile)

    # --------------------------------------------#
    #           Get the ADMIN_VIEW version        #
    #             from ambari rest api            #
    # --------------------------------------------#

    def getAdminViewVersionFromAmbari(self):
        url = self.ambariConfiguration['url'] + ":" + self.ambariConfiguration['port'] + self.ambariConfiguration['api']['getAdminViewVersion']
        r = requests.get(url, auth=(self.ambariConfiguration['user'], self.ambariConfiguration['password']), verify=False)
        data = r.json()
        if data['versions'][0]['ViewVersionInfo']['view_name'] == 'ADMIN_VIEW':
            longVersion = data['versions'][0]['ViewVersionInfo']['version']
            versionDigits = longVersion.split('.')
            self.adminViewVersion = '.'.join([versionDigits[0], versionDigits[1]])

    # --------------------------------------------#
    #           Get the queue configuration       #
    #             from ambari rest api            #
    # --------------------------------------------#

    def getQueuesFromAmbari(self, interactif=False):
        url = self.ambariConfiguration['url'] + ":" + self.ambariConfiguration['port'] + self.ambariConfiguration['api']['getQueuesFromAmbari']
        r = requests.get(url, auth=(self.ambariConfiguration['user'], self.ambariConfiguration['password']), verify=False)
        self.ambari = r.json()
        if(interactif):
            print("Retour du GET pour : " + r.url + "\nStatus : " + str(r.status_code))
            print(json.dumps(self.ambari, indent=2))

    # --------------------------------------------#
    #        Inject the queue configuration       #
    #               in ambari rest api            #
    # --------------------------------------------#

    def putQueuesInAmbari(self):
        self.getAdminViewVersionFromAmbari()
        url = self.ambariConfiguration['url'] + ":" + self.ambariConfiguration['port'] + self.ambariConfiguration['api']['putQueuesInAmbari']
        headers = defaultdict(dict)
        properties = defaultdict(dict)
        # Get actual configuration for increase the version
        self.getQueuesFromAmbari()

        # Test if those parameters are mandatory or not
        # "yarn.scheduler.capacity.root.Chats.acl_submit_applications": "*",
        # "yarn.scheduler.capacity.root.Chats.user-limit-factor": "1",
        # "yarn.scheduler.capacity.root.Chats.acl_administer_queue": "*",
        # "yarn.scheduler.capacity.root.Chats.minimum-user-limit-percent": "100",

        # Add the default configuration like root config, could be overwrited by config from the files
        for key in self.ambariConfiguration['default'].keys():
            properties[key] = self.ambariConfiguration['default'][key]
        # Iterate the Queues object by queueName to create the properties
        for queueName in sorted(self.queues.keys()):
            # if(self.configuration['root-name'] in self.queues[queueName].keys()): # and 'arborescence-head' not in self.queues[queueName].keys()):
                # We have a queue that is in the root
                # Iterate the properties for this queue
                for propertyName in sorted(self.queues[queueName].keys()):
                    # set the property if this is a property accepted in the configuration and not the root:yes property
                    if(propertyName in self.properties.keys() and propertyName != self.configuration['root-name'] and propertyName != 'arborescence-head'):
                        # Test if we are on the root queue
                        if queueName == self.configuration['root-name']:
                            properties['.'.join([self.configuration['root'], propertyName])] = self.queues[queueName][propertyName]
                        # Test if arborescence is not empty
                        elif(self.queues[queueName]['arborescence'] != ""):
                            properties['.'.join([self.configuration['root'], self.queues[queueName]['arborescence'], self.queues[queueName]['queue-name'], propertyName])] = self.queues[queueName][propertyName]
                        else:
                            properties['.'.join([self.configuration['root'], self.queues[queueName]['queue-name'], propertyName])] = self.queues[queueName][propertyName]
        desired_config = []
        desired_config.append(defaultdict(dict))
        desired_config[0]['service_config_version_note'] = self.ambariConfiguration['service_config_version_note']
        desired_config[0]['tag'] = self.ambariConfiguration['tag'] + str(int(time.time() * 1000))  # "TOPOLOGY_RESOLVED"
        desired_config[0]['type'] = "capacity-scheduler"
        desired_config[0]['version'] = self.ambari['items'][0]['version'] + 1
        desired_config[0]['properties'] = properties
        clusters = defaultdict(dict)
        clusters['desired_config'] = desired_config
        data = defaultdict(dict)
        data['Clusters'] = clusters
        print(json.dumps(data, indent=2))
        for key in self.ambariConfiguration['headers-by-version'][self.adminViewVersion]:
            headers[key] = self.ambariConfiguration['headers-by-version'][self.adminViewVersion][key]
        r = requests.put(url, headers=headers, data=json.dumps(data), auth=(self.ambariConfiguration['user'], self.ambariConfiguration['password']), verify=False)
        print("Retour du PUT pour : " + r.url + "\nStatus : " + str(r.status_code))
        print(r.text)

    # --------------------------------------------#
    #             Set the dryRun boolean           #
    # --------------------------------------------#

    def setDryRun(self, dryRun):
        self.dryRun = dryRun

    # --------------------------------------------#
    #             Set the dryRun boolean           #
    # --------------------------------------------#

    def getDryRun(self):
        return self.dryRun

    # --------------------------------------------#
    #          Show the Queues configuration      #
    # --------------------------------------------#

    def showQueues(self):
        pp = pprint.PrettyPrinter(indent=4)
        pp.pprint(dict(self.queues))

    # --------------------------------------------#
    #     Pretty print the Queues configuration   #
    # --------------------------------------------#

    def prettyPrintQueues(self):
        print(BACK_GRAY + BLACK + "\nQueues configuration" + DEFAULT + BACK_DEFAULT)
        for queueName in sorted(self.queues.keys()):
            print(BACK_BLUE + CYAN + "Queue" + DEFAULT + BACK_DEFAULT + " : " + CYAN + queueName + DEFAULT)
            for propertyName in sorted(self.queues[queueName]):
                # print('{0:{1}} {2: <20}'.format(propertyName, str(maxPropertyLenght), self.queues[queueName][propertyName]))
                print('  {: <40} {: <20}'.format(GREEN + propertyName + DEFAULT, self.queues[queueName][propertyName]))

    # --------------------------------------------#
    #    Revert the topology configuration Dict   #
    # --------------------------------------------#
    @staticmethod
    def revertConfigurationDict(configDict):
        tempoDict = {}
        for columns in configDict:
            tempoDict[configDict[columns]['property']] = {"column": columns, "default": configDict[columns]['default']}
        return tempoDict

    # --------------------------------------------#
    #       Converti la LETTRE vers code ASCII    #
    # --------------------------------------------#

    @staticmethod
    def lettreVersNum(lettre):
        return ord(lettre)

    # --------------------------------------------#
    #       Converti la LETTRE en COLONNE NUM     #
    # --------------------------------------------#

    @staticmethod
    def lettreVersCol(lettre):
        if(len(lettre) == 1):
            return (ord(lettre.upper()) - ord('A')) + 1

    # --------------------------------------------#
    #      Converti la COLONNE NUM en LETTRE      #
    # --------------------------------------------#

    @staticmethod
    def colVersLettre(column):
        if(type(column) == int):
            return chr(column + ord('A') - 1)

    # --------------------------------------------#
    #       Converti le code ASCII en LETTRE      #
    # --------------------------------------------#

    @staticmethod
    def numVersLettre(chiffre):
        if (chiffre >= ord('a') and chiffre <= ord('z')) or (chiffre >= ord('A') and chiffre <= ord('Z')):
            return chr(chiffre)

    # --------------------------------------------#
    #               Read the XLS file             #
    # --------------------------------------------#

    def readXlsFile(self, fileXLS, configXLS):
        print(BACK_GRAY + BLACK + "\nReading XLS file :" + DEFAULT + BACK_DEFAULT + " " + fileXLS)
        wb = load_workbook(fileXLS, data_only=True)
        ws = wb.get_sheet_by_name(configXLS['sheet-name'])
        row = int(configXLS['cellule-origine']['row'])
        col = int(configXLS['cellule-origine']['col'])
        rowMax = int(configXLS['row-max'])
        # Check the column titles if they are same as the configuration
        for column in configXLS['topology']:
            if(ws[column + configXLS['row-titles']].value != configXLS['topology'][column]['columnTitle']):
                # Error on a column chack
                print(RED + "ERROR : column analyse of the excel file is incoherent in the cell " + column + configXLS['row-titles'] +
                      " find : " + ws[column + configXLS['row-titles']].value + " instead of : " + configXLS['topology'][column]['columnTitle'] + DEFAULT)
                return False
        # parsing file
        actualArborescence = ""
        while row != rowMax:
            if(ws.cell(row=row, column=col).value is None and ws.cell(row=row, column=col + 1).value is None):
                # Empty line, arborescence is reinitialized
                actualArborescence = ""
            else:
                if(ws.cell(row=row, column=col).value is not None and ws.cell(row=row, column=int(configXLS['queues-name-column'])).value is None):
                    # New arborescence
                    # actualArborescence = ws.cell(row=row, column=col).value
                    # We add the arborescence-head key in the queue
                    # self.addQueueValue(actualArborescence, actualArborescence, self.configuration['arbo-queues-property'], '')
                    # self.addQueueValue(actualArborescence, actualArborescence, 'arborescence-head', 'yes')
                    queueName = ws.cell(row=row, column=col).value
                    # self.addQueueValue(actualArborescence, queueName, self.configuration['arbo-queues-property'], '')
                    self.addQueueValue(actualArborescence, queueName, 'arborescence-head', 'yes')
                    print(BACK_BLUE + CYAN + "New ARBORESCENCE : " + actualArborescence + DEFAULT + BACK_DEFAULT)
                    for column in sorted(configXLS['topology']):
                        # Store the cell value
                        cellValue = str(ws.cell(row=row, column=self.lettreVersCol(column)).value)
                        # Check that we are not in the Queue column and on the right of the Queue column
                        if(column.upper() != self.colVersLettre(int(configXLS['queues-name-column'])) and column.upper() > self.colVersLettre(int(configXLS['queues-name-column']))):
                            # Add the value if present, in other case, the default value if in configuration file
                            if(cellValue != 'None'):
                                self.addQueueValue(actualArborescence, queueName, configXLS['topology'][column]['property'], cellValue)
                                sys.stdout.write(GREEN + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                            # default value
                            elif(configXLS['topology'][column]['default'] != ""):
                                self.addQueueValue(actualArborescence, queueName, configXLS['topology'][column]['property'], configXLS['topology'][column]['default'])
                                sys.stdout.write(YELLOW + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                            # no default value and cell empty, do nothing
                            else:
                                sys.stdout.write(RED + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                        # Queue column
                        elif(column.upper() == self.colVersLettre(int(configXLS['queues-name-column']))):
                            if(actualArborescence != ""):
                                sys.stdout.write("Arborescence : " + CYAN + actualArborescence + "." + queueName + DEFAULT + ", ")
                            else:
                                sys.stdout.write("Arborescence : " + CYAN + queueName + DEFAULT + ", ")
                        print("")
                    actualArborescence = queueName
                # if(ws.cell(row=row, column=int(configXLS['queues-name-column'])).value is not None):
                elif(ws.cell(row=row, column=int(configXLS['queues-name-column'])).value is not None):
                    # Find a Queue
                    queueName = str(ws.cell(row=row, column=int(configXLS['queues-name-column'])).value)
                    # Add the root key for this queue
                    self.addQueueValue(actualArborescence, queueName, self.configuration['root-name'], 'yes')
                    # Iterate on the line with the confiured columns
                    for column in sorted(configXLS['topology']):
                        # Store the cell value
                        cellValue = str(ws.cell(row=row, column=self.lettreVersCol(column)).value)
                        # Check that we are not in the Queue column and on the right of the Queue column
                        if(column.upper() != self.colVersLettre(int(configXLS['queues-name-column'])) and column.upper() > self.colVersLettre(int(configXLS['queues-name-column']))):
                            # Add the value if present, in other case, the default value if in configuration file
                            if(cellValue != 'None'):
                                self.addQueueValue(actualArborescence, queueName, configXLS['topology'][column]['property'], cellValue)
                                sys.stdout.write(GREEN + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                            # default value
                            elif(configXLS['topology'][column]['default'] != ""):
                                self.addQueueValue(actualArborescence, queueName, configXLS['topology'][column]['property'], configXLS['topology'][column]['default'])
                                sys.stdout.write(YELLOW + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                            # no default value and cell empty, do nothing
                            else:
                                sys.stdout.write(RED + configXLS['topology'][column]['property'] + ": " + cellValue + ", " + DEFAULT)
                        # Queue column
                        elif(column.upper() == self.colVersLettre(int(configXLS['queues-name-column']))):
                            if(actualArborescence != ""):
                                sys.stdout.write("Arborescence : " + CYAN + actualArborescence + "." + queueName + DEFAULT + ", ")
                            else:
                                sys.stdout.write("Arborescence : " + CYAN + queueName + DEFAULT + ", ")
                        print("")
            row += 1
        # We construct the tree of leafs
        self.manageQueuesTreeLeafs()
        # Do some checks on the queues
        return self.checkQueuesCoherence()

    # --------------------------------------------#
    #               Read the XML file             #
    # --------------------------------------------#

    def readXmlFile(self, fileXML, configXML, configProperties):
        print(BACK_GRAY + BLACK + "\nReading XML file :" + DEFAULT + BACK_DEFAULT + " " + fileXML)
        # Regular expression compilation for root or preroot
        regPreRoot = re.compile(r"^" + re.escape(self.configuration['pre-root']))
        regRoot = re.compile(r"^" + re.escape(self.configuration['root']))
        tree = etree.parse(fileXML)
        # Iterate on the XML file data
        for prop in tree.iter('property'):
            # Split the string in elements
            # Compute number of elements in the key
            # yarn.scheduler.capacity.root.SROM.capacity
            # 1   .2        .3       .4   .5   .6
            elements = prop.find('name').text.split('.')
            nbElements = len(elements)
            if(regRoot.match(prop.find('name').text) is not None):
                # Find a property with the root
                if(str(elements[-1]) not in configProperties.keys()):
                    # We find a property that is not in  the configuration file properties-config part
                    print(BACK_YELLOW + BLACK + "Property not in properties-config section :" + DEFAULT + BACK_DEFAULT + " " + str(elements[-1]) +
                          " for : " + prop.find('name').text)
                else:
                    if(nbElements > 5):
                        # Add the queue with its arborescence and the value
                        # We join the arborescence elements, extract the queue, the property and the value
                        self.addQueueValue('.'.join(elements[4:-2]), str(elements[-2]), str(elements[-1]), prop.find('value').text)
                        # We add the fact that this is a root arborescence
                        self.addQueueValue('.'.join(elements[4:-2]), str(elements[-2]), self.configuration['root-name'], 'yes')
                    else:
                        # Add to root the general config
                        self.addQueueValue("", str(elements[-2]), str(elements[-1]), prop.find('value').text)
            elif(regPreRoot.match(prop.find('name').text) is not None):
                # Find a property without the root, add to the general configuration Queue
                if(nbElements == 4):
                    self.addQueueValue("", str(self.configuration['general']), str(elements[-1]), prop.find('value').text)
                # Find a property with a particular structure
                # ex : yarn.scheduler.capacity.queue-mappings-override.enable : false
                # make a dict
                elif(nbElements == 5):
                    self.addQueueValue("", str(self.configuration['general']), str(elements[-2]), {str(elements[-1]): prop.find('value').text})
                else:
                    print(BACK_RED + BLACK + "Property not traited :" + DEFAULT + BACK_DEFAULT + " " + prop.find('name').text)
            else:
                print(BACK_RED + BLACK + "Property not traited :" + DEFAULT + BACK_DEFAULT + " " + prop.find('name').text)

# --------------------------------------------#
#                   Code                      #
# --------------------------------------------#

# --------------------------------------------#
#               Show the version              #
# --------------------------------------------#


def programVersion():
    print("Version : 0.0.1")


# --------------------------------------------#
#               Exit with Error               #
# --------------------------------------------#

def exitWithError(errorText):
    print("\nERROR : " + errorText + "\n")
    sys.exit(1)


# --------------------------------------------#
#           Read the configuration            #
#                  JSON file                  #
# --------------------------------------------#

def fileReaderJSON(fileName):
    global vg_configuration
    global vg_configProperties
    global vg_xlsConfig
    global vg_ambariConfig
    with open(fileName) as jsonFile:
        jsonData = json.load(jsonFile)
    vg_configuration = jsonData['configuration']
    vg_configProperties = jsonData['properties-config']
    vg_xlsConfig = jsonData['xls-config']
    vg_ambariConfig = jsonData['ambari-config']


# --------------------------------------------#
#             Command line parsing            #
# --------------------------------------------#

def parseCommandLine():
    validFromArgument = False
    validToArgument = False
    global vg_arguments
    parser = argparse.ArgumentParser(
        description='Yarn Queue Manager for setting or reading queues configuration', prog='YarnQueueManager')
    parser.add_argument('-v', '--version', action='store_true', default=False, help='print the version')
    parser.add_argument('-V', '--verbose', action='store_true', default=False, help='verbose mode')
    parser.add_argument('-p', '--print', action='store_true', default=False, help='print configuration')
    parser.add_argument('-d', '--dryRun', action='store_true', default=False, help='Dry run only, nothing is modified')
    parser.add_argument('-f', '--from', type=str, help='Get capacity-scheduler configuration from [ambari|xlsFile|xmlFile|jsonFile]')
    parser.add_argument('-t', '--to', type=str, help='Put capacity-scheduler configuration to [ambari|xlsFile|jsonFile]')
    parser.add_argument('-e', '--xlsFile', type=str, help='Excel file name for get or put')
    parser.add_argument('-j', '--jsonFile', type=str, help='Json file name for get or put')
    parser.add_argument('-x', '--xmlFile', type=str, help='Xml file name for get ex : capacity-scheduler.xml')

    vg_arguments = vars(parser.parse_args())
    print(vg_arguments)
    fileReaderJSON('conf/YarnQueueManager.json')
    # Initiate the object
    queues = Queues(vg_configuration, vg_configProperties, vg_ambariConfig)
    # Set the check boolean for knowing if we will change something or not
    queues.setDryRun(vg_arguments['dryRun'])

    if vg_arguments['version']:
        programVersion()
        sys.exit

    # Get the source configuration from...
    if vg_arguments['from'] is not None:
        # Get AMBARI configuration
        if vg_arguments['from'] == 'ambari':
            queues.getQueuesFromAmbari(True)

        # Get EXCEL FILE configuration
        elif vg_arguments['from'] == 'xlsFile':
            if vg_arguments['xlsFile'] is not None:
                if not queues.readXlsFile(vg_arguments['xlsFile'], vg_xlsConfig):
                    exitWithError("Error detected when reading excel file.")
            else:
                print("Arguments : \n" + str(vg_arguments))
                exitWithError("Invalid arguments : when using <from xlsFile> you have to set <xlsFile> parameter.")

        # Get XML FILE configuration aka : capacity-scheduler.xml
        elif vg_arguments['from'] == 'xmlFile':
            if vg_arguments['xmlFile'] is not None:
                queues.readXmlFile(vg_arguments['xmlFile'], vg_xlsConfig, vg_configProperties)
            else:
                print("Arguments : \n" + str(vg_arguments))
                exitWithError("Invalid arguments : when using <from xmlFile> you have to set <xmlFile> parameter.")

        # Get JSON FILE configuration
        elif vg_arguments['from'] == 'jsonFile':
            if vg_arguments['jsonFile'] is not None:
                queues.readJsonFile(vg_arguments['jsonFile'])
            else:
                print("Arguments : \n" + str(vg_arguments))
                exitWithError("Invalid arguments : when using <from jsonFile> you have to set <xlsFile> parameter.")
        else:
            print("Arguments : \n" + str(vg_arguments))
            exitWithError("Invalid arguments : you have not set a valid <from [ambari|xlsFile|jsonFile]> parameter.")
        validFromArgument = True
    else:
        print("Arguments : \n" + str(vg_arguments))
        exitWithError("Invalid arguments : you have not set a <from [ambari|xlsFile|jsonFile]> parameter.")

    # Want to print the loaded configuration
    if vg_arguments['print'] is not None:
        queues.prettyPrintQueues()


    # Put the configuration to...
    if vg_arguments['to'] is not None and validFromArgument and queues.getDryRun() is False:
        # Put the configuration to AMBARI
        if vg_arguments['to'] == 'ambari':
            queues.putQueuesInAmbari()
        # Put the configuration to EXCEL FILE
        elif vg_arguments['to'] == 'xlsFile':
            if vg_arguments['xlsFile'] is not None:
                queues.queuesToXLS(vg_arguments['xlsFile'], vg_xlsConfig)
            else:
                print("Arguments : \n" + str(vg_arguments))
                exitWithError("Invalid arguments : when using <to xlsFile> you have to set <xlsFile> parameter.")
        # Put the configuration to JSON FILE
        elif vg_arguments['to'] == 'jsonFile':
            if vg_arguments['jsonFile'] is not None:
                queues.queuesToJsonFile(vg_arguments['jsonFile'])
            else:
                print("Arguments : \n" + str(vg_arguments))
                exitWithError("Invalid arguments : when using <to jsonFile> you have to set <xlsFile> parameter.")
        else:
            print("Arguments : \n" + str(vg_arguments))
            exitWithError("Invalid arguments : you have not set a valid <to [ambari|xlsFile|jsonFile]> parameter.")
        validToArgument = True
    # Only shwing the configuration from...
    elif queues.getDryRun() is True or vg_arguments['print'] is not None:
        validToArgument = True
    else:
        print("Arguments : \n" + str(vg_arguments))
        exitWithError("Invalid arguments : you have not set a <to [ambari|xlsFile|jsonFile]> parameter.")

    if validFromArgument is False or validToArgument is False:
        print("Arguments : \n" + str(vg_arguments))
        exitWithError("Invalid arguments.")

    # if vg_arguments['xml'] is None and vg_arguments['excel'] is not None:
    #     # Read the XLS File and show the queues
    #     queues.readXlsFile(vg_arguments['excel'], vg_xlsConfig)
    #     queues.prettyPrintQueues()
    # elif vg_arguments['xml'] is not None and vg_arguments['excel'] is not None:
    #     # Read the XML file with the actual configuration and generate the XLS file
    #     queues.readXmlFile(vg_arguments['xml'], vg_xlsConfig, vg_configProperties)
    #     queues.queuesToXLS(vg_arguments['excel'], vg_xlsConfig)
    #     queues.prettyPrintQueues()
    # elif vg_arguments['list'] is not None:
    #     # Read the XML file with the actual configuration and generate the XLS file
    #     actual_config = queues.getQueuesFromAmbari(True)
    #     if(vg_arguments['save'] is not None):
    #         queues.saveQueuesToFile(actual_config, vg_arguments['save'])
    # else:
    #     print("Arguments : \n" + str(vg_arguments))
    #     exitWithError("Invalid arguments : file and delimiter must be defined.")

    # if vg_arguments['ambari']:
    #     # Option to send to Ambari the configuration
    #     queues.putQueuesInAmbari()


# --------------------------------------------#
#                     Main                    #
# --------------------------------------------#

def main():
    parseCommandLine()


if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding('utf8')
    main()
